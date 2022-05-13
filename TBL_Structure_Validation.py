from sys import builtin_module_names
import pandas as pd
import openpyxl as excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Fill, Color
import os

from pymysql import ROWID

#TBL_COLS1_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\BIA_DEV_INFO.csv"
#TBL_COLS2_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\BIA_TST_INFO.csv"

TBL_COLS1_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\Snowflake\TST_CLM_CYQ_SGNTR_TEMP.csv"
TBL_COLS2_csv = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\Teradata\TST_V1_CLM_CYQ_SGNTR.csv"

#fDtlDiffs = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\TBL_COL_Diffs.csv"
#fDtlDiffsXLSX = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\Snowflake\Compare\TBL_COL_Diffs.xlsx"
fDtlDiffs = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\CLM_CYQ_SGNTR_Diffs.csv"
fDtlDiffsXLSX = r"C:\Users\user\OneDrive - Apprio, Inc\Documents\PBAR\Historical\CLM_CYQ_SGNTR_Diffs.xlsx"


def removeDFCols(df, lstCols2Remove):
    # sColTxt = Column to drop from Data Frame

    lstCols = df.columns.values.tolist()

    for sColTxt in lstCols2Remove:
        idx = lstCols.count(sColTxt)
        if idx > 0:
            df.drop(sColTxt, axis=1, inplace=True)

    return df    


def removeIDRNulls(df):

    lstCols = df.columns.values.tolist()
    for col in lstCols:
        df.loc[df[col] == "?", col] = '' 

    return df


def sortDF(df):

    lstCols = df.columns.values.tolist()
    #df = df.sort_values(by=['TABLE_SCHEMA', 'TABLE_NAME', 'COLUMN_NAME'])
    df = df.sort_values(by=lstCols)
    
    return df 


def main():
    #########################################################
    # Define variables
    #########################################################

    # truncate output file if exists
    if os.path.exists(fDtlDiffs):
        os.truncate(fDtlDiffs, 0)

    #########################################################
    # Read CME and IDR csv files into Pandas Data Frames
    #########################################################
    dfFile1 = pd.read_csv(TBL_COLS1_csv, dtype=str, na_filter=False) 
    dfFile1 = removeDFCols(dfFile1, ["IDR_INSRT_TS", "IDR_UPDT_TS"]) 
    dfFile1 = removeIDRNulls(dfFile1)
    ##dfFile1.sort_values(by=['TABLE_SCHEMA', 'TABLE_NAME', 'COLUMN_NAME'], inplace=True)

    dfFile2 = pd.read_csv(TBL_COLS2_csv, dtype=str, na_filter=False) 
    dfFile2 = removeDFCols(dfFile2, ["IDR_INSRT_TS", "IDR_UPDT_TS"]) 
    dfFile2 = removeIDRNulls(dfFile2)
    ##dfFile2.sort_values(by=['TABLE_SCHEMA', 'TABLE_NAME', 'COLUMN_NAME'], inplace=True)

    print("NOF dfFile1 rows:"+str(len(dfFile1.index)))
    print("NOF dfFile2 rows:"+str(len(dfFile2.index)))

    #########################################################################################
    # in essence a join/comparison on all columns
    # NOTE: Setting "indicator=True" adds a column to the merged DataFame where the value 
    #       of each row can be one of three possible values: left_only, right_only, or both
    # NOTE2: Another option "first_df.compare(second_df, keep_equal=True)"
    #########################################################################################
    comparison_df = dfFile1.merge(dfFile2, indicator=True, how='outer')
    comparison_df = sortDF(comparison_df)
    comparison_df.to_csv(fDtlDiffs,sep = ",", index=False, line_terminator = "\n")  

    # after merge --> row will be "marked" as 'both" if both data frame rows match
    dfDiff = comparison_df[comparison_df['_merge'] != 'both']
    dfDiff = sortDF(dfDiff)

    exit(0)
    # sort and write out results    
    
    ##dfDiffSorted.to_csv(fDtlDiffs,sep = ",", index=False, line_terminator = "\n")
    dfDiff.to_csv(fDtlDiffs,sep = ",", index=False, line_terminator = "\n")    
    
    exit(0)

    # saving diffs to xlsx file
    XLSX = pd.ExcelWriter(fDtlDiffsXLSX)
    dfDiffSorted.to_excel(XLSX, index=False)    
    XLSX.save()


    ###\/######
    # Add hilighting of rows
    wrkbk = excel.load_workbook(fDtlDiffsXLSX)

    if wrkbk is None:
        print("couldn't get workbook onject")
    else:
        print("Got the workbook")    

    ###################################################
    # process Excel spreadsheet
    ###################################################
    bFlipColor = False

    sheet = wrkbk.active
    rowColor = "FFFF00"

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

        if row[0].row == 1:
            pass
            # skip header
        elif (row[0].row) % 2 == 0:
            if bFlipColor == False:
                bFlipColor = True
                rowColor = 'B0FFFF'
            else:
                bFlipColor = False
                rowColor = '0DFFBE'  


        for cell in row:
            if cell.row == 1:
                # color header row
                cell.fill = PatternFill(start_color='66DDFF', end_color='66DDFF', fill_type = "solid")                
            else:
                cell.fill = PatternFill(start_color=rowColor, end_color=rowColor, fill_type = "solid")
                                        
    wrkbk.save(fDtlDiffsXLSX)

  
    exit(0)


if __name__ == "__main__":
    
    main()

